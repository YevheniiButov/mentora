"""
Admin Users Export Routes
Handles filtering, searching, and XLSX export of user data for partner organizations.
"""
import io
import logging
from datetime import datetime, timezone

from flask import Blueprint, render_template, request, send_file, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy import or_

from extensions import db
from models import User

logger = logging.getLogger(__name__)

admin_export_bp = Blueprint('admin_export', __name__, url_prefix='/admin/users')


# ─── Human-readable column headers for the Excel export ───────────────────────
EXPORT_COLUMNS = {
    'first_name':          'First Name',
    'last_name':           'Last Name',
    'email':               'Email',
    'phone':               'Phone',
    'city':                'City',
    'profession':          'Profession',
    'specialization':      'Specialization',
    'university_name':     'University',
    'dutch_level':         'Dutch Proficiency',
    'english_level':       'English Proficiency',
    'legal_status':        'Legal Status',
    'diploma_status':      'Diploma Status',
    'big_exam_registered': 'BIG Exam Status',
    'exam_date':           'Planned Exam Date',
    'housing_needed':      'Housing Needed',
    'work_as_assistant':   'Ready to Work as Assistant',
    'created_at':          'Registration Date',
}


# ─── Helper: build filtered query from request args ───────────────────────────
def _build_query(args):
    """Build SQLAlchemy query with optional filters from URL query params."""
    query = User.query.filter(
        User.is_deleted == False,
        User.role == 'user',
    )

    if profession := args.get('profession'):
        query = query.filter(User.profession == profession)

    if dutch_level := args.get('dutch_level'):
        query = query.filter(User.dutch_level == dutch_level)

    if legal_status := args.get('legal_status'):
        query = query.filter(User.legal_status == legal_status)

    if housing_needed := args.get('housing_needed'):
        query = query.filter(User.housing_needed == housing_needed)

    if big_exam := args.get('big_exam_registered'):
        query = query.filter(User.big_exam_registered == big_exam)

    if diploma_status := args.get('diploma_status'):
        query = query.filter(User.diploma_status == diploma_status)

    # Free-text search across name / email / city
    if search := args.get('search', '').strip():
        like = f'%{search}%'
        query = query.filter(or_(
            User.first_name.ilike(like),
            User.last_name.ilike(like),
            User.email.ilike(like),
            User.city.ilike(like),
        ))

    return query


# ─── Page: filter UI ──────────────────────────────────────────────────────────
@admin_export_bp.route('/export', methods=['GET'])
@login_required
def export_page():
    """Render the admin export / filter page."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Forbidden'}), 403

    args = request.args
    query = _build_query(args)

    # Pagination
    page = args.get('page', 1, type=int)
    per_page = 25
    pagination = query.order_by(User.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Unique values for filter dropdowns
    def distinct_values(col):
        rows = db.session.query(col).filter(col.isnot(None)).distinct().all()
        return sorted([r[0] for r in rows if r[0]])

    filter_options = {
        'professions':    distinct_values(User.profession),
        'dutch_levels':   ['A1', 'A2', 'B1', 'B2', 'C1', 'C2', 'native'],
        'legal_statuses': distinct_values(User.legal_status),
        'housing_opts':   distinct_values(User.housing_needed),
        'big_exam_opts':  distinct_values(User.big_exam_registered),
        'diploma_opts':   distinct_values(User.diploma_status),
    }

    return render_template(
        'admin/users_export_panel.html',
        users=pagination.items,
        pagination=pagination,
        filter_options=filter_options,
        active_filters=args,
        total_filtered=pagination.total,
    )


# ─── Download: XLSX ───────────────────────────────────────────────────────────
@admin_export_bp.route('/export/download', methods=['GET'])
@login_required
def export_download():
    """Generate and stream a partner-ready XLSX file."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Forbidden'}), 403

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500

    try:
        users = _build_query(request.args).order_by(User.created_at.desc()).all()

        # ── Build workbook ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Candidates'

        # Styles
        header_font  = Font(name='Calibri', bold=True, color='FFFFFFFF', size=11)
        header_fill  = PatternFill('solid', fgColor='FF1A3A5C')
        accent_fill  = PatternFill('solid', fgColor='FFE8F0FB')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align   = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border  = Border(
            bottom=Side(style='thin', color='FFD0D7E3'),
            right=Side(style='thin', color='FFD0D7E3'),
        )

        col_keys = list(EXPORT_COLUMNS.keys())
        headers  = list(EXPORT_COLUMNS.values())

        # Write header row
        ws.row_dimensions[1].height = 30
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

        # Write data rows
        for row_idx, user in enumerate(users, start=2):
            ws.row_dimensions[row_idx].height = 18
            for col_idx, key in enumerate(col_keys, start=1):
                value = getattr(user, key, None)
                # Format dates
                if value is not None and hasattr(value, 'strftime'):
                    value = value.strftime('%d.%m.%Y')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = left_align
                cell.border    = thin_border
                if row_idx % 2 == 0:
                    cell.fill = accent_fill

        # Simple column width setting
        for col_idx in range(1, len(col_keys) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # Freeze header row
        ws.freeze_panes = 'A2'

        # ── Metadata sheet ────────────────────────────────────────────────────
        meta_ws = wb.create_sheet('Export Info')
        meta_ws['A1'] = 'Export generated by'
        meta_ws['B1'] = f'{current_user.first_name or ""} {current_user.last_name or ""} (ID: {current_user.id})'
        meta_ws['A2'] = 'Export date'
        meta_ws['B2'] = datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M UTC')
        meta_ws['A3'] = 'Total candidates'
        meta_ws['B3'] = len(users)
        meta_ws['A4'] = 'Platform'
        meta_ws['B4'] = 'Mentora — bigmentor.nl'

        # ── Write to buffer and build response safely ─────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'mentora_candidates_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

        # ── Audit log ─────────────────────────────────────────────────────────
        logger.info(
            'ADMIN_EXPORT | admin_id=%s email=%s exported=%d filters=%s',
            current_user.id,
            current_user.email,
            len(users),
            dict(request.args),
        )

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error('ADMIN_EXPORT ERROR: %s', str(e), exc_info=True)
        return jsonify({'error': f'Export failed: {str(e)}'}), 500
